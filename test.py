from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side

# Read the sales data from the Excel file into a DataFrame
sales_modified_df = pd.read_excel(
    "report_data/sales_modified.xlsx",
    sheet_name="Sales_export",
    engine="openpyxl",
)

# Get the unique brands from the DataFrame
unique_brands = sales_modified_df["Brand"].unique().tolist()
length_of_unique_portal = (
        len(sales_modified_df["Portal"].unique().tolist()) + 1
)  # +1 for (Total) Row

# Initialize an empty DataFrame to store the combined pivot tables
combined_pivot_table = pd.DataFrame()

# Get the current time and create a DataFrame with it
current_time = datetime.now().strftime("%H:%M")
time_column = pd.DataFrame({"Current Time": [current_time]})

# Initialize border style
border_style = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

# Create a heading row DataFrame with column names
heading_row = pd.DataFrame(
    {
        "Brand": ["Brand"],
        "Current Time": ["Current Time"],
        "Portal": ["Portal"],
        "Total Order Qty": ["Total Order Qty"],
        "Total Order Amount (in Lacs)": ["Total Order Amount (in Lacs)"],
        "ASP": ["ASP"]
    }
)

# Create an empty DataFrame with three rows
empty_rows = pd.DataFrame({"": [""]}, index=range(1))

# * ----------------    Samsonite and Samsonite Red Pivot    ----------------
# Filter data for "SAMSONITE"
samsonite_df = sales_modified_df[
    (sales_modified_df["Brand"] == "SAMSONITE")
    & (~sales_modified_df["SKU Desc"].str.contains("SAMSONITE RED"))
    ]
if samsonite_df.empty:
    # Create a dummy table with 0 values
    columns = [
        "Portal",
        "Total Order Qty",
        "Total Order Amount (in Lacs)",
        "ASP"
    ]
    dummy_data = pd.DataFrame([["ALL", 0, 0, 0]], columns=columns)
    samsonite_pivot_table = dummy_data
else:
    # Create a pivot table for "SAMSONITE"
    samsonite_pivot_table = pd.pivot_table(
        samsonite_df,
        index="Portal",
        values=["Order Qty", "Order Amount"],
        aggfunc={"Order Qty": "sum", "Order Amount": "sum"},
    ).reset_index()

    # Rename columns in the "SAMSONITE" pivot table
    samsonite_pivot_table = samsonite_pivot_table.rename(
        columns={
            "Order Qty": "Total Order Qty",
            "Order Amount": "Total Order Amount (in Lacs)",
        }
    )

# Filter data for "SAMSONITE RED"
samsonite_red_df = sales_modified_df[
    (sales_modified_df["Brand"] == "SAMSONITE")
    & (sales_modified_df["SKU Desc"].str.contains("SAMSONITE RED"))
    ]
if samsonite_red_df.empty:
    # Create a dummy table with 0 values
    columns = [
        "Portal",
        "Total Order Qty",
        "Total Order Amount (in Lacs)",
        "ASP"
    ]
    dummy_data = pd.DataFrame([["ALL", 0, 0, 0]], columns=columns)
    samsonite_red_pivot_table = dummy_data
else:
    # Create a pivot table for "SAMSONITE"
    samsonite_red_pivot_table = pd.pivot_table(
        samsonite_red_df,
        index="Portal",
        values=["Order Qty", "Order Amount"],
        aggfunc={"Order Qty": "sum", "Order Amount": "sum"},
    ).reset_index()

    # Rename columns in the "SAMSONITE" pivot table
    samsonite_red_pivot_table = samsonite_red_pivot_table.rename(
        columns={
            "Order Qty": "Total Order Qty",
            "Order Amount": "Total Order Amount (in Lacs)",
        }
    )

# * ----------------    AT Backpack, AT Trolly bag and AT Duffle bag Pivot    ----------------
# Filter data for "AT Backpack"
at_backpack_df = sales_modified_df[
    (sales_modified_df["Brand"] == "AMERICAN TOURISTER")
    & (sales_modified_df["SKU Desc"].str.contains("BACKPACK"))
    ]
if at_backpack_df.empty:
    # Create a dummy table with 0 values
    columns = [
        "Portal",
        "Total Order Qty",
        "Total Order Amount (in Lacs)",
        "ASP"
    ]
    dummy_data = pd.DataFrame([["ALL", 0, 0, 0]], columns=columns)
    at_backpack_pivot_table = dummy_data
else:
    # Create a pivot table for "SAMSONITE"
    at_backpack_pivot_table = pd.pivot_table(
        at_backpack_df,
        index="Portal",
        values=["Order Qty", "Order Amount"],
        aggfunc={"Order Qty": "sum", "Order Amount": "sum"},
    ).reset_index()

    # Rename columns in the "SAMSONITE" pivot table
    at_backpack_pivot_table = at_backpack_pivot_table.rename(
        columns={
            "Order Qty": "Total Order Qty",
            "Order Amount": "Total Order Amount (in Lacs)",
        }
    )

# Filter data for "AT Trolly Bag"
at_trolly_bag_df = sales_modified_df[
    (sales_modified_df["Brand"] == "AMERICAN TOURISTER")
    & (sales_modified_df["SKU Desc"].str.contains("TROLLY BAG"))
    ]
if at_trolly_bag_df.empty:
    # Create a dummy table with 0 values
    columns = [
        "Portal",
        "Total Order Qty",
        "Total Order Amount (in Lacs)",
        "ASP"
    ]
    dummy_data = pd.DataFrame([["ALL", 0, 0, 0]], columns=columns)
    at_trolly_bag_pivot_table = dummy_data
else:
    # Create a pivot table for "SAMSONITE"
    at_trolly_bag_pivot_table = pd.pivot_table(
        at_trolly_bag_df,
        index="Portal",
        values=["Order Qty", "Order Amount"],
        aggfunc={"Order Qty": "sum", "Order Amount": "sum"},
    ).reset_index()

    # Rename columns in the "SAMSONITE" pivot table
    at_trolly_bag_pivot_table = at_trolly_bag_pivot_table.rename(
        columns={
            "Order Qty": "Total Order Qty",
            "Order Amount": "Total Order Amount (in Lacs)",
        }
    )

# Filter data for "AT Duffle Bag"
at_duffle_bag_df = sales_modified_df[
    (sales_modified_df["Brand"] == "AMERICAN TOURISTER")
    & (sales_modified_df["SKU Desc"].str.contains("DUFFLE BAG"))
    ]
if at_duffle_bag_df.empty:
    # Create a dummy table with 0 values
    columns = [
        "Portal",
        "Total Order Qty",
        "Total Order Amount (in Lacs)",
        "ASP"
    ]
    dummy_data = pd.DataFrame([["ALL", 0, 0, 0]], columns=columns)
    at_duffle_bag_pivot_table = dummy_data
else:
    # Create a pivot table for "SAMSONITE"
    at_duffle_bag_pivot_table = pd.pivot_table(
        at_duffle_bag_df,
        index="Portal",
        values=["Order Qty", "Order Amount"],
        aggfunc={"Order Qty": "sum", "Order Amount": "sum"},
    ).reset_index()

    # Rename columns in the "SAMSONITE" pivot table
    at_duffle_bag_pivot_table = at_duffle_bag_pivot_table.rename(
        columns={
            "Order Qty": "Total Order Qty",
            "Order Amount": "Total Order Amount (in Lacs)",
        }
    )

# ! ---------------  (Total) Row for SAMSONITE  ---------------
# Calculate Total Order Qty and Total Order Amount for SAMSONITE
samsonite_total_order_qty = samsonite_pivot_table["Total Order Qty"].sum()
samsonite_total_order_amount = samsonite_pivot_table[
    "Total Order Amount (in Lacs)"
].sum()
samsonite_pivot_table["ASP"] = (
            samsonite_pivot_table["Total Order Amount (in Lacs)"] / samsonite_pivot_table["Total Order Qty"]).round(0)
asp = samsonite_total_order_amount / samsonite_total_order_qty
asp = round(asp)
samsonite_pivot_table["Total Order Amount (in Lacs)"] = (
        samsonite_pivot_table["Total Order Amount (in Lacs)"] / 100000
).round(2)
# Create a DataFrame for the Total row for SAMSONITE
samsonite_total_row = pd.DataFrame(
    {
        "Portal": ["Total"],
        "Total Order Qty": [samsonite_total_order_qty],
        "Total Order Amount (in Lacs)": [samsonite_total_order_amount],
        "ASP": [asp]
    }
)

# Concatenate the Total row with the SAMSONITE pivot table
samsonite_pivot_table = pd.concat(
    [samsonite_pivot_table, samsonite_total_row], ignore_index=True
)

# ! ---------------  (Total) Row for SAMSONITE RED  ---------------
# Calculate Total Order Qty and Total Order Amount for SAMSONITE RED
samsonite_red_total_order_qty = samsonite_red_pivot_table["Total Order Qty"].sum()
samsonite_red_total_order_amount = samsonite_red_pivot_table[
    "Total Order Amount (in Lacs)"
].sum()
samsonite_red_pivot_table["ASP"] = (
            samsonite_red_pivot_table["Total Order Amount (in Lacs)"] / samsonite_red_pivot_table[
        "Total Order Qty"]).round(0)
asp = samsonite_red_total_order_amount / samsonite_red_total_order_qty
asp = round(asp)
samsonite_red_pivot_table["Total Order Amount (in Lacs)"] = (
        samsonite_red_pivot_table["Total Order Amount (in Lacs)"] / 100000
).round(2)
# Create a DataFrame for the Total row for SAMSONITE RED
samsonite_red_total_row = pd.DataFrame(
    {
        "Portal": ["Total"],
        "Total Order Qty": [samsonite_red_total_order_qty],
        "Total Order Amount (in Lacs)": [samsonite_red_total_order_amount],
        "ASP": [asp]
    }
)

# Concatenate the Total row with the SAMSONITE RED pivot table
samsonite_red_pivot_table = pd.concat(
    [samsonite_red_pivot_table, samsonite_red_total_row], ignore_index=True
)

# ! ---------------  (Total) Row for AT BACKPACK  ---------------
# Calculate Total Order Qty and Total Order Amount for SKECHERS APPARELS
at_backpack_total_order_qty = at_backpack_pivot_table["Total Order Qty"].sum()
at_backpack_total_order_amount = at_backpack_pivot_table[
    "Total Order Amount (in Lacs)"
].sum()
at_backpack_pivot_table["ASP"] = (
            at_backpack_pivot_table["Total Order Amount (in Lacs)"] / at_backpack_pivot_table["Total Order Qty"]).round(
    0)
asp = at_backpack_total_order_amount / at_backpack_total_order_qty
asp = round(asp)
at_backpack_pivot_table["Total Order Amount (in Lacs)"] = (
        at_backpack_pivot_table["Total Order Amount (in Lacs)"] / 100000
).round(2)
# Create a DataFrame for the Total row for SKECHERS APPARELS
at_backpack_total_row = pd.DataFrame(
    {
        "Portal": ["Total"],
        "Total Order Qty": [at_backpack_total_order_qty],
        "Total Order Amount (in Lacs)": [at_backpack_total_order_amount],
        "ASP": [asp]
    }
)

# Concatenate the Total row with the SKECHERS APPARELS pivot table
at_backpack_pivot_table = pd.concat(
    [at_backpack_pivot_table, at_backpack_total_row], ignore_index=True
)

# ! ---------------  (Total) Row for AT TROLLY BAG  ---------------
# Calculate Total Order Qty and Total Order Amount for SKECHERS APPARELS
at_trolly_bag_total_order_qty = at_trolly_bag_pivot_table["Total Order Qty"].sum()
at_trolly_bag_total_order_amount = at_trolly_bag_pivot_table[
    "Total Order Amount (in Lacs)"
].sum()
at_trolly_bag_pivot_table["ASP"] = (
            at_trolly_bag_pivot_table["Total Order Amount (in Lacs)"] / at_trolly_bag_pivot_table[
        "Total Order Qty"]).round(0)
asp = at_trolly_bag_total_order_amount / at_trolly_bag_total_order_qty
asp = round(asp)
at_trolly_bag_pivot_table["Total Order Amount (in Lacs)"] = (
        at_trolly_bag_pivot_table["Total Order Amount (in Lacs)"] / 100000
).round(2)
# Create a DataFrame for the Total row for SKECHERS APPARELS
at_trolly_bag_total_row = pd.DataFrame(
    {
        "Portal": ["Total"],
        "Total Order Qty": [at_trolly_bag_total_order_qty],
        "Total Order Amount (in Lacs)": [at_trolly_bag_total_order_amount],
        "ASP": [asp]
    }
)

# Concatenate the Total row with the SKECHERS APPARELS pivot table
at_trolly_bag_pivot_table = pd.concat(
    [at_trolly_bag_pivot_table, at_trolly_bag_total_row], ignore_index=True
)

# ! ---------------  (Total) Row for AT DUFFLE BAG  ---------------
# Calculate Total Order Qty and Total Order Amount for SKECHERS APPARELS
at_duffle_bag_total_order_qty = at_duffle_bag_pivot_table["Total Order Qty"].sum()
at_duffle_bag_total_order_amount = at_duffle_bag_pivot_table[
    "Total Order Amount (in Lacs)"
].sum()
at_duffle_bag_pivot_table["ASP"] = (at_duffle_bag_pivot_table["Total Order Amount (in Lacs)"] / at_duffle_bag_pivot_table["Total Order Qty"]).round(0)
asp = at_duffle_bag_total_order_amount / at_duffle_bag_total_order_qty
print(asp, at_duffle_bag_total_order_qty, at_duffle_bag_total_order_amount)
if not pd.isna(asp):
    asp = round(asp)
at_duffle_bag_pivot_table["Total Order Amount (in Lacs)"] = (
        at_duffle_bag_pivot_table["Total Order Amount (in Lacs)"] / 100000
).round(2)
# Create a DataFrame for the Total row for SKECHERS APPARELS
at_duffle_bag_total_row = pd.DataFrame(
    {
        "Portal": ["Total"],
        "Total Order Qty": [at_duffle_bag_total_order_qty],
        "Total Order Amount (in Lacs)": [at_duffle_bag_total_order_amount],
        "ASP": [asp]
    }
)

# Concatenate the Total row with the SKECHERS APPARELS pivot table
at_duffle_bag_pivot_table = pd.concat(
    [at_duffle_bag_pivot_table, at_duffle_bag_total_row], ignore_index=True
)

# ! "SAMSONITE"
# Create brand and time columns for "SKECHERS KIDSWEAR" pivot table
samsonite_brand_column = pd.DataFrame({"Brand": ["SAMSONITE"]})
samsonite_time_column = pd.DataFrame({"Current Time": [current_time]})

# Concatenate brand, time, and pivot table dataframes for "SKECHERS APPARELS"
samsonite_first_conc = pd.concat(
    [samsonite_brand_column, samsonite_time_column], axis=1
)
samsonite_second_conc = pd.concat([samsonite_first_conc, samsonite_pivot_table], axis=1)

# Concatenate heading row, concatenated dataframes, and empty rows vertically for "SKECHERS APPARELS"
samsonite_brand_time_pivot = pd.concat(
    [heading_row, samsonite_second_conc, empty_rows], axis=0
)

# ! "SAMSONITE RED"
# Create brand and time columns for "SKECHERS KIDSWEAR" pivot table
samsonite_red_brand_column = pd.DataFrame({"Brand": ["SAMSONITE RED"]})
samsonite_red_time_column = pd.DataFrame({"Current Time": [current_time]})

# Concatenate brand, time, and pivot table dataframes for "SKECHERS APPARELS"
samsonite_red_first_conc = pd.concat(
    [samsonite_red_brand_column, samsonite_red_time_column], axis=1
)
samsonite_red_second_conc = pd.concat(
    [samsonite_red_first_conc, samsonite_red_pivot_table], axis=1
)

# Concatenate heading row, concatenated dataframes, and empty rows vertically for "SKECHERS APPARELS"
samsonite_red_brand_time_pivot = pd.concat(
    [heading_row, samsonite_red_second_conc, empty_rows], axis=0
)

# ! "AMERICAN TOURISTER BACKPACK"
# Create brand and time columns for "SKECHERS KIDSWEAR" pivot table
at_backpack_brand_column = pd.DataFrame({"Brand": ["AT BACKPACK"]})
at_backpack_time_column = pd.DataFrame({"Current Time": [current_time]})

# Concatenate brand, time, and pivot table dataframes for "SKECHERS APPARELS"
at_backpack_first_conc = pd.concat(
    [at_backpack_brand_column, at_backpack_time_column], axis=1
)
at_backpack_second_conc = pd.concat(
    [at_backpack_first_conc, at_backpack_pivot_table], axis=1
)

# Concatenate heading row, concatenated dataframes, and empty rows vertically for "SKECHERS APPARELS"
at_backpack_brand_time_pivot = pd.concat(
    [heading_row, at_backpack_second_conc, empty_rows], axis=0
)

# ! "AMERICAN TOURISTER TROLLY BAG"
# Create brand and time columns for "SKECHERS KIDSWEAR" pivot table
at_trolly_bag_brand_column = pd.DataFrame({"Brand": ["AT TROLLY BAG"]})
at_trolly_bag_time_column = pd.DataFrame({"Current Time": [current_time]})

# Concatenate brand, time, and pivot table dataframes for "SKECHERS APPARELS"
at_trolly_bag_first_conc = pd.concat(
    [at_trolly_bag_brand_column, at_trolly_bag_time_column], axis=1
)
at_trolly_bag_second_conc = pd.concat(
    [at_trolly_bag_first_conc, at_trolly_bag_pivot_table], axis=1
)

# Concatenate heading row, concatenated dataframes, and empty rows vertically for "SKECHERS APPARELS"
at_trolly_bag_brand_time_pivot = pd.concat(
    [heading_row, at_trolly_bag_second_conc, empty_rows], axis=0
)

# ! "AMERICAN TOURISTER DUFFLE BAG"
# Create brand and time columns for "SKECHERS KIDSWEAR" pivot table
at_duffle_bag_brand_column = pd.DataFrame({"Brand": ["AT DUFFLE BAG"]})
at_duffle_bag_time_column = pd.DataFrame({"Current Time": [current_time]})

# Concatenate brand, time, and pivot table dataframes for "SKECHERS APPARELS"
at_duffle_bag_first_conc = pd.concat(
    [at_duffle_bag_brand_column, at_duffle_bag_time_column], axis=1
)
at_duffle_bag_second_conc = pd.concat(
    [at_duffle_bag_first_conc, at_duffle_bag_pivot_table], axis=1
)

# Concatenate heading row, concatenated dataframes, and empty rows vertically for "SKECHERS APPARELS"
at_duffle_bag_brand_time_pivot = pd.concat(
    [heading_row, at_duffle_bag_second_conc, empty_rows], axis=0
)

# Concatenate the pivot tables for "SKECHERS" and "SKECHERS APPARELS" with the combined pivot table
combined_pivot_table = pd.concat(
    [
        combined_pivot_table,
        samsonite_brand_time_pivot,
        samsonite_red_brand_time_pivot,
        at_backpack_brand_time_pivot,
        at_trolly_bag_brand_time_pivot,
        at_duffle_bag_brand_time_pivot,
    ],
    ignore_index=True,
)

# * ----------------    End of Brand - Location Pivot    ----------------

# Reset the index of the combined pivot table
combined_pivot_table.reset_index(drop=True, inplace=True)

# Write the combined pivot table to an Excel file
with pd.ExcelWriter(
        "report_data/MIS_report.xlsx",
        engine="openpyxl",
        mode="w",
) as writer:
    # Remove the existing "For MIS" sheet if it exists
    if "For MIS" in writer.book.sheetnames:
        writer.book.remove(writer.book["For MIS"])  # Remove existing sheet if it exists

    combined_pivot_table.to_excel(
        writer,
        sheet_name="For MIS",
        index=False,
    )

workbook = load_workbook("report_data/MIS_report.xlsx")
sheet = workbook.active

# Remove the first row
sheet.delete_rows(1)

# Define the headers to be bolded
headers_to_bold = [
    "Brand",
    "Current Time",
    "Portal",
    "Total Order Qty",
    "Total Order Amount (in Lacs)",
    "ASP"
]

# Iterate over the rows in the sheet (up to row 60)
for row in sheet.iter_rows(min_row=1, max_row=60):
    for cell in row:
        # Check if the value in the cell matches any of the headers to be bolded
        if cell.value in headers_to_bold:
            cell.font = Font(bold=True)

# * Iterate C column check on every row where Portal is located in Bold get the row number
# * ----------    Brands Border    ----------
start_lst = []
for start, row in enumerate(sheet.iter_rows(min_col=3, max_col=3), start=1):
    # Check if the font of the cell in the "C" column is bold
    if row[0].value == "Portal" and row[0].font.bold:
        # If the "C" column is in bold, add the row number to the list
        start_lst.append(start)

end_lst = []
for end, row in enumerate(sheet.iter_rows(min_col=3, max_col=3), start=1):
    if "Total Order Amount (in Lacs)" in str(row[0].value):
        break
    # Check if the cell in the "C" column is empty
    if row[0].value is None or (
            isinstance(row[0].value, str) and row[0].value.strip() == ""
    ):
        end_lst.append(end - 1)

pairs = [(start_lst[i], end_lst[i]) for i in range(len(start_lst))]

for pair in pairs:
    start_row, end_row = pair
    for row in sheet.iter_rows(
            min_row=start_row, max_row=end_row, min_col=1, max_col=5
    ):
        for cell in row:
            cell.border = border_style

# * ----------  Auto-fit the columns  ----------
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Adjusting the width for padding
    sheet.column_dimensions[column].width = adjusted_width

# Iterate through each row in column C
for row in sheet.iter_rows(min_row=1, min_col=3, max_col=3):
    for cell in row:
        if cell.value == "Total":
            # Apply bold font to each cell in the row
            for c in range(1, sheet.max_column + 1):
                sheet.cell(row=cell.row, column=c).font = openpyxl.styles.Font(
                    bold=True
                )
workbook.save("report_data/MIS_report.xlsx")
